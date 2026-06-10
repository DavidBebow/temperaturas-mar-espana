"""
obtener_embalses.py
===================
Descarga el Excel semanal oficial del Boletín Hidrológico del MITECO,
cruza los datos con el diccionario de embalses por provincia y genera:
  - docs/embalses/{provincia}.json  (un fichero por provincia)
  - docs/embalses_nacional.json     (resumen nacional para el mapa de España)

Fuente: Boletín Hidrológico Semanal — MITECO (datos oficiales)
URL patrón: https://sede.miteco.gob.es/BoleHWeb/accion/cargador_archivo.htm
            ?file=cache/xls/{AAAAMM}/{AAAASSNN}_es.xls

Ejecución: python scripts/obtener_embalses.py
Dependencias: pip install requests openpyxl
"""

import requests
import json
import os
import io
from datetime import datetime, date, timedelta

# ── UTILIDADES ─────────────────────────────────────────────────────────────────

def semana_iso(d: date):
    """Devuelve (año_iso, semana_iso) para una fecha."""
    iso = d.isocalendar()
    return iso[0], iso[1]

def url_boletin(anyo: int, semana: int) -> str:
    """
    Construye la URL del Excel del boletín para un año y semana ISO dados.
    El MITECO usa el patrón:
        cache/xls/AAAAMM/AAAASSNN_es.xls
    donde MM = mes del lunes de esa semana ISO, SS = semana con 2 dígitos,
    NN = número de boletín del año (= semana con 2 dígitos, por lo general igual a SS).
    """
    # Calcular el lunes de la semana ISO pedida
    lunes = date.fromisocalendar(anyo, semana, 1)
    mes   = lunes.strftime("%m")
    ss    = f"{semana:02d}"
    base  = f"https://sede.miteco.gob.es/BoleHWeb/accion/cargador_archivo.htm"
    ruta  = f"cache/xls/{anyo}{mes}/{anyo}{ss}{ss}_es.xls"
    return f"{base}?file={ruta}&mimetype=application/vnd.ms-excel"

def descargar_excel(max_semanas_atras: int = 4):
    """
    Intenta descargar el Excel del boletín más reciente.
    Prueba desde la semana actual hacia atrás hasta max_semanas_atras.
    Devuelve (bytes_del_fichero, fecha_lunes) o (None, None).
    """
    hoy = date.today()
    for delta in range(max_semanas_atras):
        d = hoy - timedelta(weeks=delta)
        anyo, semana = semana_iso(d)
        url = url_boletin(anyo, semana)
        lunes = date.fromisocalendar(anyo, semana, 1)
        print(f"  Intentando semana {semana}/{anyo} ({lunes.strftime('%d/%m/%Y')}) → {url[:80]}...")
        try:
            r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
            if r.status_code == 200 and len(r.content) > 10_000:
                print(f"  ✓ Descargado ({len(r.content)//1024} KB)")
                return r.content, lunes
            else:
                print(f"    HTTP {r.status_code}, tamaño {len(r.content)} bytes — no válido")
        except Exception as e:
            print(f"    Error: {e}")
    return None, None

def leer_excel(contenido: bytes) -> dict:
    """
    Lee el Excel del MITECO y devuelve un dict:
        { nombre_embalse_lower: {"vol": float, "pct": float} }
    La hoja principal se llama normalmente 'Embalses' o similar.
    Columnas relevantes: nombre, volumen actual (hm³), % llenado.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)

    # Buscar la hoja correcta (puede llamarse 'Embalses', 'Sheet1', etc.)
    hoja = None
    for nombre_hoja in wb.sheetnames:
        print(f"  Hoja encontrada: '{nombre_hoja}'")
        if any(k in nombre_hoja.lower() for k in ["embalse", "reserva", "sheet", "datos"]):
            hoja = wb[nombre_hoja]
            break
    if hoja is None:
        hoja = wb.active

    print(f"  Usando hoja: '{hoja.title}'")

    datos = {}
    col_nombre = col_vol = col_pct = None

    for fila in hoja.iter_rows(values_only=True):
        # Detectar cabecera buscando palabras clave
        if col_nombre is None:
            fila_lower = [str(c).lower() if c else "" for c in fila]
            for i, celda in enumerate(fila_lower):
                if "embalse" in celda or "nombre" in celda or "presa" in celda:
                    col_nombre = i
                if "volumen" in celda and ("actual" in celda or "reserva" in celda or col_vol is None):
                    col_vol = i
                if "%" in celda or "porcent" in celda or "llenado" in celda:
                    col_pct = i
            if col_nombre is not None:
                print(f"  Cabecera detectada → nombre={col_nombre}, vol={col_vol}, pct={col_pct}")
            continue

        nombre = fila[col_nombre] if col_nombre is not None and len(fila) > col_nombre else None
        if not nombre or not isinstance(nombre, str) or len(nombre.strip()) < 2:
            continue

        vol = None
        pct = None
        try:
            if col_vol is not None and len(fila) > col_vol:
                v = fila[col_vol]
                if isinstance(v, (int, float)) and v > 0:
                    vol = float(v)
            if col_pct is not None and len(fila) > col_pct:
                p = fila[col_pct]
                if isinstance(p, (int, float)) and 0 <= p <= 150:
                    pct = float(p)
        except Exception:
            pass

        # Si tenemos vol pero no pct, calculamos pct más adelante con la capacidad del diccionario
        clave = nombre.strip().lower()
        datos[clave] = {"vol": vol, "pct": pct}

    print(f"  Total embalses leídos del Excel: {len(datos)}")
    return datos

# ── DICCIONARIO MAESTRO ────────────────────────────────────────────────────────
# Para cada embalse se define:
#   buscar: lista de términos a buscar en el nombre del Excel (minúsculas, sin acentos)
#   lat/lon: coordenadas del centro del embalse
#   cap: capacidad total en hm³ (fuente: estadoembalses.es / MITECO)
#   rio, municipio: metadatos para el popup del mapa

def norm(s):
    """Normaliza texto para comparación: minúsculas, sin acentos."""
    import unicodedata
    s = s.lower().strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")

PROVINCIAS = {
    # ══ ANDALUCÍA ══
    "cadiz": {
        "nombre": "Cádiz", "comunidad": "Andalucía",
        "embalses": [
            {"id":"celemin",       "nombre":"Celemín",        "buscar":["celemi"],            "rio":"Celemín",     "municipio":"Los Barrios",       "cap":44.8,  "lat":36.167,"lon":-5.620},
            {"id":"guadalcacin",   "nombre":"Guadalcacín",    "buscar":["guadalcacin"],       "rio":"Majaceite",   "municipio":"Jerez de la Frontera","cap":800.3,"lat":36.720,"lon":-5.830},
            {"id":"arcos",         "nombre":"Arcos",          "buscar":["arcos"],             "rio":"Guadalete",   "municipio":"Arcos de la Frontera","cap":14.6, "lat":36.745,"lon":-5.797},
            {"id":"guadarranque",  "nombre":"Guadarranque",   "buscar":["guadarranque"],      "rio":"Guadarranque","municipio":"San Roque",          "cap":87.7,  "lat":36.232,"lon":-5.462},
            {"id":"barbate",       "nombre":"Barbate",        "buscar":["barbate"],           "rio":"Barbate",     "municipio":"Vejer de la Frontera","cap":228.1,"lat":36.253,"lon":-5.830},
            {"id":"los_hurones",   "nombre":"Los Hurones",    "buscar":["huron"],             "rio":"Majaceite",   "municipio":"Grazalema",          "cap":135.3, "lat":36.760,"lon":-5.590},
            {"id":"bornos",        "nombre":"Bornos",         "buscar":["bornos"],            "rio":"Guadalete",   "municipio":"Bornos",             "cap":200.2, "lat":36.803,"lon":-5.715},
            {"id":"charco_redondo","nombre":"Charco Redondo", "buscar":["charco redondo"],    "rio":"Palmones",    "municipio":"Los Barrios",        "cap":81.6,  "lat":36.220,"lon":-5.617},
            {"id":"almodovar",     "nombre":"Almodóvar",      "buscar":["almodovar"],         "rio":"Bugones",     "municipio":"Castellar",          "cap":5.7,   "lat":36.310,"lon":-5.440},
            {"id":"zahara",        "nombre":"Zahara-El Gastor","buscar":["zahara"],           "rio":"Guadalete",   "municipio":"Zahara de la Sierra", "cap":222.7, "lat":36.837,"lon":-5.428},
        ]
    },
    "malaga": {
        "nombre": "Málaga", "comunidad": "Andalucía",
        "embalses": [
            {"id":"guadalteba",    "nombre":"Guadalteba",     "buscar":["guadalteba"],        "rio":"Guadalhorce", "municipio":"Ardales",            "cap":154.0, "lat":36.887,"lon":-4.885},
            {"id":"guadalhorce_e", "nombre":"Guadalhorce",    "buscar":["guadalhorce"],       "rio":"Guadalhorce", "municipio":"Ardales",            "cap":125.7, "lat":36.862,"lon":-4.857},
            {"id":"la_vinuela",    "nombre":"La Viñuela",     "buscar":["vinuela","viñuela"],  "rio":"Guaro",      "municipio":"La Viñuela",         "cap":165.4, "lat":36.871,"lon":-4.149},
            {"id":"concepcion",    "nombre":"Concepción",     "buscar":["concepcion"],        "rio":"Verde",       "municipio":"Marbella",           "cap":61.9,  "lat":36.590,"lon":-4.990},
            {"id":"limonero",      "nombre":"Limonero",       "buscar":["limonero"],          "rio":"Guadalmedina","municipio":"Málaga",             "cap":22.3,  "lat":36.756,"lon":-4.437},
            {"id":"casasola",      "nombre":"Casasola",       "buscar":["casasola"],          "rio":"Campanillas", "municipio":"Málaga",             "cap":21.7,  "lat":36.726,"lon":-4.614},
            {"id":"conde_guadalh", "nombre":"Conde Guadalhorce","buscar":["conde guadal"],    "rio":"Turón",       "municipio":"Ardales",            "cap":125.7, "lat":36.864,"lon":-4.793},
            {"id":"montejaque",    "nombre":"Montejaque",     "buscar":["montejaque"],        "rio":"Guadarés",    "municipio":"Montejaque",         "cap":36.0,  "lat":36.726,"lon":-5.280},
        ]
    },
    "almeria": {
        "nombre": "Almería", "comunidad": "Andalucía",
        "embalses": [
            {"id":"beninar",       "nombre":"Benínar",        "buscar":["beninar","benínar"],  "rio":"Adra",       "municipio":"Berja",              "cap":61.7,  "lat":36.882,"lon":-2.982},
            {"id":"cuevas",        "nombre":"Cuevas Almanzora","buscar":["cuevas"],           "rio":"Almanzora",   "municipio":"Cuevas del Almanzora","cap":161.3,"lat":37.328,"lon":-1.884},
        ]
    },
    "huelva": {
        "nombre": "Huelva", "comunidad": "Andalucía",
        "embalses": [
            {"id":"olivargas",     "nombre":"Olivargas",      "buscar":["olivargas"],         "rio":"Olivargas",   "municipio":"Calañas",            "cap":29.0,  "lat":37.658,"lon":-6.882},
            {"id":"corumbel",      "nombre":"Corumbel Bajo",  "buscar":["corumbel"],          "rio":"Corumbel",    "municipio":"Beas",               "cap":18.0,  "lat":37.384,"lon":-6.922},
            {"id":"zufre",         "nombre":"Zufre",          "buscar":["zufre"],             "rio":"Huelva",      "municipio":"Zufre",              "cap":175.3, "lat":37.836,"lon":-6.334},
            {"id":"aracena",       "nombre":"Aracena",        "buscar":["aracena"],           "rio":"Huelva",      "municipio":"Aracena",            "cap":128.0, "lat":37.879,"lon":-6.604},
            {"id":"piedras",       "nombre":"Piedras",        "buscar":["piedras"],           "rio":"Piedras",     "municipio":"El Almendro",        "cap":59.5,  "lat":37.560,"lon":-7.272},
            {"id":"andevalo",      "nombre":"Andévalo",       "buscar":["andevalo","andévalo"],"rio":"Malagón",    "municipio":"El Granado",         "cap":634.4, "lat":37.674,"lon":-7.090},
            {"id":"odiel",         "nombre":"Odiel",          "buscar":["odiel"],             "rio":"Odiel",       "municipio":"Valverde del Camino", "cap":8.0,   "lat":37.614,"lon":-6.769},
            {"id":"chanza",        "nombre":"Chanza",         "buscar":["chanza"],            "rio":"Chanza",      "municipio":"Paymogo",            "cap":341.4, "lat":37.820,"lon":-7.426},
            {"id":"jarrama",       "nombre":"Jarrama",        "buscar":["jarrama"],           "rio":"Jarrama",     "municipio":"Jabugo",             "cap":42.6,  "lat":37.909,"lon":-6.787},
            {"id":"sancho",        "nombre":"Sancho",         "buscar":["sancho"],            "rio":"Meca",        "municipio":"Huelva",             "cap":58.0,  "lat":37.267,"lon":-6.967},
            {"id":"cala",          "nombre":"Cala",           "buscar":["cala"],              "rio":"Cala",        "municipio":"Cala",               "cap":58.8,  "lat":37.990,"lon":-6.280},
        ]
    },
    "granada": {
        "nombre": "Granada", "comunidad": "Andalucía",
        "embalses": [
            {"id":"cubillas",      "nombre":"Cubillas",       "buscar":["cubillas"],          "rio":"Cubillas",    "municipio":"Iznalloz",           "cap":13.5,  "lat":37.481,"lon":-3.634},
            {"id":"colomera",      "nombre":"Colomera",       "buscar":["colomera"],          "rio":"Colomera",    "municipio":"Colomera",           "cap":40.2,  "lat":37.415,"lon":-3.637},
            {"id":"el_portillo",   "nombre":"El Portillo",    "buscar":["portillo"],          "rio":"Castril",     "municipio":"Castril",            "cap":31.3,  "lat":37.817,"lon":-2.805},
            {"id":"quentar",       "nombre":"Quéntar",        "buscar":["quentar","quéntar"], "rio":"Aguas Blancas","municipio":"Quéntar",           "cap":13.6,  "lat":37.196,"lon":-3.569},
            {"id":"rules",         "nombre":"Rules",          "buscar":["rules"],             "rio":"Guadalfeo",   "municipio":"Vélez de Benaudalla","cap":113.3, "lat":36.814,"lon":-3.573},
            {"id":"canales",       "nombre":"Canales",        "buscar":["canales"],           "rio":"Genil",       "municipio":"Güéjar Sierra",      "cap":70.0,  "lat":37.138,"lon":-3.558},
            {"id":"los_bermejales","nombre":"Los Bermejales", "buscar":["bermejal"],          "rio":"Cacín",       "municipio":"Arenas del Rey",     "cap":91.0,  "lat":36.987,"lon":-4.044},
            {"id":"beznar",        "nombre":"Béznar",         "buscar":["beznar","béznar"],   "rio":"Ízbor",       "municipio":"El Pinar",           "cap":52.9,  "lat":36.915,"lon":-3.548},
            {"id":"negratin",      "nombre":"Negratín",       "buscar":["negratin","negratín"],"rio":"Guadiana Menor","municipio":"Freila",          "cap":571.0, "lat":37.656,"lon":-2.981},
            {"id":"fco_abellan",   "nombre":"Francisco Abellán","buscar":["abellan","abellán"],"rio":"Fardes",     "municipio":"Guadix",             "cap":58.1,  "lat":37.394,"lon":-3.125},
            {"id":"san_clemente",  "nombre":"San Clemente",   "buscar":["san clemente"],      "rio":"Guardal",     "municipio":"Huéscar",            "cap":117.0, "lat":37.892,"lon":-2.701},
        ]
    },
    "sevilla": {
        "nombre": "Sevilla", "comunidad": "Andalucía",
        "embalses": [
            {"id":"el_pintado",    "nombre":"El Pintado",     "buscar":["pintado"],           "rio":"Viar",        "municipio":"El Real de la Jara", "cap":215.0, "lat":37.803,"lon":-5.913},
            {"id":"los_melonares", "nombre":"Los Melonares",  "buscar":["melonares"],         "rio":"Viar",        "municipio":"Guillena",           "cap":185.6, "lat":37.721,"lon":-6.041},
            {"id":"el_agrio",      "nombre":"El Agrio",       "buscar":["agrio"],             "rio":"Agrio",       "municipio":"Aznalcóllar",        "cap":20.3,  "lat":37.534,"lon":-6.278},
            {"id":"jose_toran",    "nombre":"José Torán",     "buscar":["toran","torán"],     "rio":"Guadalbacar", "municipio":"Morón de la Frontera","cap":113.2,"lat":37.148,"lon":-5.422},
            {"id":"huesna",        "nombre":"Huesna",         "buscar":["huesna"],            "rio":"Huesna",      "municipio":"San Nicolás del Puerto","cap":134.6,"lat":37.979,"lon":-5.739},
            {"id":"gergal",        "nombre":"El Gergal",      "buscar":["gergal"],            "rio":"Huelva",      "municipio":"Guillena",           "cap":35.0,  "lat":37.665,"lon":-6.025},
            {"id":"la_puebla_caz", "nombre":"La Puebla de Cazalla","buscar":["puebla de cazalla"],"rio":"Corbones","municipio":"La Puebla de Cazalla","cap":73.7,"lat":37.200,"lon":-5.230},
            {"id":"la_minilla",    "nombre":"La Minilla",     "buscar":["minilla"],           "rio":"Huelva",      "municipio":"Real de la Jara",    "cap":58.0,  "lat":37.630,"lon":-5.860},
            {"id":"torre_aguila",  "nombre":"La Torre del Águila","buscar":["torre del aguila","aguila"],"rio":"Salado","municipio":"Utrera",        "cap":64.4,  "lat":37.040,"lon":-5.680},
        ]
    },
    "cordoba": {
        "nombre": "Córdoba", "comunidad": "Andalucía",
        "embalses": [
            {"id":"arenoso",       "nombre":"Arenoso",        "buscar":["arenoso"],           "rio":"Arenoso",     "municipio":"Hornachuelos",       "cap":167.0, "lat":37.860,"lon":-5.320},
            {"id":"yeguas",        "nombre":"Yeguas",         "buscar":["yeguas"],            "rio":"Yeguas",      "municipio":"Montoro",            "cap":230.0, "lat":38.000,"lon":-4.447},
            {"id":"bembezar",      "nombre":"Bembézar",       "buscar":["bembezar","bembézar"],"rio":"Bembézar",   "municipio":"Hornachuelos",       "cap":328.0, "lat":37.834,"lon":-5.247},
            {"id":"san_rafael_nav","nombre":"San Rafael Navallana","buscar":["navallana","san rafael de navallana"],"rio":"Guadalmellato","municipio":"Córdoba","cap":160.1,"lat":37.901,"lon":-4.830},
            {"id":"puente_nuevo",  "nombre":"Puente Nuevo",   "buscar":["puente nuevo"],      "rio":"Guadiato",    "municipio":"Espiel",             "cap":282.0, "lat":38.053,"lon":-5.101},
            {"id":"guadalmellato", "nombre":"Guadalmellato",  "buscar":["guadalmellato"],     "rio":"Guadalmellato","municipio":"Córdoba",           "cap":149.0, "lat":37.990,"lon":-4.775},
            {"id":"vadomojon",     "nombre":"Vadomojón",      "buscar":["vadomojon","vadomojón"],"rio":"Guadajoz", "municipio":"Baena",              "cap":163.2, "lat":37.670,"lon":-4.271},
            {"id":"la_colada",     "nombre":"La Colada",      "buscar":["la colada"],         "rio":"Guadamatilla","municipio":"Villanueva de Madrid","cap":58.0, "lat":38.250,"lon":-4.770},
            {"id":"iznajar",       "nombre":"Iznájar",        "buscar":["iznajar","iznájar"], "rio":"Genil",       "municipio":"Iznájar",            "cap":981.0, "lat":37.267,"lon":-4.310},
            {"id":"la_breña_ii",   "nombre":"La Breña II",    "buscar":["breña"],             "rio":"Guadiato",    "municipio":"Hornachuelos",       "cap":823.0, "lat":37.885,"lon":-5.164},
        ]
    },
    "jaen": {
        "nombre": "Jaén", "comunidad": "Andalucía",
        "embalses": [
            {"id":"quiebrajano",   "nombre":"Quiebrajano",    "buscar":["quiebrajano"],       "rio":"Quiebrajano", "municipio":"Jaén",               "cap":31.6,  "lat":37.795,"lon":-3.745},
            {"id":"guadalmena",    "nombre":"Guadalmena",     "buscar":["guadalmena"],        "rio":"Guadalmena",  "municipio":"Villanueva del Arzobispo","cap":346.5,"lat":38.375,"lon":-2.936},
            {"id":"el_tranco",     "nombre":"El Tranco de Beas","buscar":["tranco"],          "rio":"Guadalquivir","municipio":"Hornos",             "cap":505.7, "lat":38.039,"lon":-2.803},
            {"id":"jandula",       "nombre":"Jándula",        "buscar":["jandula","jándula"], "rio":"Jándula",     "municipio":"Andújar",            "cap":325.1, "lat":38.177,"lon":-4.100},
            {"id":"guadalen",      "nombre":"Guadalén",       "buscar":["guadalen","guadalén"],"rio":"Guadalén",   "municipio":"Vilches",            "cap":162.6, "lat":38.228,"lon":-3.553},
            {"id":"giribaile",     "nombre":"Giribaile",      "buscar":["giribaile"],         "rio":"Guadalimar",  "municipio":"Vilches",            "cap":491.1, "lat":38.087,"lon":-3.613},
            {"id":"la_fernandina", "nombre":"La Fernandina",  "buscar":["fernandina"],        "rio":"Guarrizas",   "municipio":"La Carolina",        "cap":246.0, "lat":38.226,"lon":-3.796},
            {"id":"rumblar",       "nombre":"Rumblar",        "buscar":["rumblar"],           "rio":"Rumblar",     "municipio":"Baños de la Encina", "cap":126.0, "lat":38.228,"lon":-3.796},
            {"id":"la_bolera",     "nombre":"La Bolera",      "buscar":["bolera"],            "rio":"Guadalentín", "municipio":"Pozo Alcón",         "cap":54.0,  "lat":37.824,"lon":-2.928},
        ]
    },
    # ══ MURCIA ══
    "murcia": {
        "nombre": "Murcia", "comunidad": "Región de Murcia",
        "embalses": [
            {"id":"cenajo",        "nombre":"El Cenajo",      "buscar":["cenajo"],            "rio":"Segura",      "municipio":"Hellín",             "cap":437.5, "lat":38.347,"lon":-1.688},
            {"id":"alfonso_xiii",  "nombre":"Alfonso XIII",   "buscar":["alfonso xiii"],      "rio":"Segura",      "municipio":"Lorca",              "cap":70.0,  "lat":38.214,"lon":-1.728},
            {"id":"la_cierva",     "nombre":"La Cierva",      "buscar":["cierva"],            "rio":"Segura",      "municipio":"Moratalla",          "cap":12.0,  "lat":38.075,"lon":-1.592},
            {"id":"valdeinfierno", "nombre":"Valdeinfierno",  "buscar":["valdeinfierno"],     "rio":"Luchena",     "municipio":"Lorca",              "cap":11.3,  "lat":37.953,"lon":-1.872},
            {"id":"puentes",       "nombre":"Puentes",        "buscar":["puentes"],           "rio":"Guadalentín", "municipio":"Lorca",              "cap":45.3,  "lat":37.776,"lon":-1.787},
            {"id":"argos",         "nombre":"Argos",          "buscar":["argos"],             "rio":"Argos",       "municipio":"Moratalla",          "cap":11.3,  "lat":38.338,"lon":-1.907},
            {"id":"santomera",     "nombre":"Santomera",      "buscar":["santomera"],         "rio":"Rambla Salada","municipio":"Santomera",         "cap":17.9,  "lat":38.072,"lon":-1.057},
        ]
    },
    # ══ CLM ══
    "albacete": {
        "nombre": "Albacete", "comunidad": "Castilla-La Mancha",
        "embalses": [
            {"id":"fuensanta",     "nombre":"Fuensanta",      "buscar":["fuensanta"],         "rio":"Segura",      "municipio":"Yeste",              "cap":210.0, "lat":38.334,"lon":-2.115},
            {"id":"talave",        "nombre":"Talave",         "buscar":["talave"],            "rio":"Mundo",       "municipio":"Liétor",             "cap":34.9,  "lat":38.373,"lon":-2.130},
            {"id":"camarillas",    "nombre":"Camarillas",     "buscar":["camarillas"],        "rio":"Mundo",       "municipio":"Isso (Hellín)",      "cap":36.0,  "lat":38.440,"lon":-1.886},
            {"id":"taibilla",      "nombre":"Taibilla",       "buscar":["taibilla"],          "rio":"Taibilla",    "municipio":"Nerpio",             "cap":9.0,   "lat":38.185,"lon":-2.274},
        ]
    },
    "ciudad_real": {
        "nombre": "Ciudad Real", "comunidad": "Castilla-La Mancha",
        "embalses": [
            {"id":"torre_abraham", "nombre":"Torre de Abraham","buscar":["torre de abraham"], "rio":"Bullaque",    "municipio":"Piedrabuena",        "cap":184.0, "lat":39.030,"lon":-4.422},
            {"id":"montoro_cr",    "nombre":"Montoro",        "buscar":["montoro"],           "rio":"Montoro",     "municipio":"Montoro de la Sierra","cap":105.0,"lat":38.680,"lon":-3.547},
            {"id":"vicario",       "nombre":"Vicario",        "buscar":["vicario"],           "rio":"Guadiana",    "municipio":"Almagro",            "cap":32.9,  "lat":38.826,"lon":-3.839},
            {"id":"gasset",        "nombre":"Gasset",         "buscar":["gasset"],            "rio":"Becea",       "municipio":"Ciudad Real",        "cap":41.7,  "lat":38.932,"lon":-3.784},
            {"id":"penarroya_cr",  "nombre":"Peñarroya",      "buscar":["peñarroya cr","penarr"],"rio":"Guadiana", "municipio":"Argamasilla de Alba","cap":50.3,  "lat":38.800,"lon":-4.067},
        ]
    },
    "toledo": {
        "nombre": "Toledo", "comunidad": "Castilla-La Mancha",
        "embalses": [
            {"id":"castrejon",     "nombre":"Castrejón",      "buscar":["castrejon","castrejón"],"rio":"Tajo",     "municipio":"La Puebla de Montalbán","cap":42.0,"lat":39.810,"lon":-4.465},
            {"id":"rosarito",      "nombre":"Rosarito",       "buscar":["rosarito"],          "rio":"Tiétar",      "municipio":"Candeleda",          "cap":86.0,  "lat":40.049,"lon":-5.243},
            {"id":"guajaraz",      "nombre":"Guajaraz",       "buscar":["guajaraz"],          "rio":"Guajaraz",    "municipio":"Argés",              "cap":18.1,  "lat":39.770,"lon":-4.100},
            {"id":"azutan",        "nombre":"Azután",         "buscar":["azutan","azután"],   "rio":"Tajo",        "municipio":"Azután",             "cap":84.0,  "lat":39.791,"lon":-5.143},
            {"id":"navalcan",      "nombre":"Navalcán",       "buscar":["navalcan","navalcán"],"rio":"Porquerizo", "municipio":"Navalcán",           "cap":39.0,  "lat":40.000,"lon":-5.127},
            {"id":"finisterre",    "nombre":"Finisterre",     "buscar":["finisterre"],        "rio":"Algodor",     "municipio":"Tembleque",          "cap":133.0, "lat":39.590,"lon":-3.430},
            {"id":"el_castro",     "nombre":"El Castro",      "buscar":["castro to","el castro algodor"],"rio":"Algodor","municipio":"Villamuelas","cap":27.3,   "lat":39.820,"lon":-3.750},
        ]
    },
    "cuenca": {
        "nombre": "Cuenca", "comunidad": "Castilla-La Mancha",
        "embalses": [
            {"id":"contreras",     "nombre":"Contreras",      "buscar":["contreras"],         "rio":"Cabriel",     "municipio":"Contreras",          "cap":361.0, "lat":39.540,"lon":-1.481},
            {"id":"alarcon",       "nombre":"Alarcón",        "buscar":["alarcon","alarcón"],  "rio":"Júcar",      "municipio":"Alarcón",            "cap":1112.0,"lat":39.554,"lon":-2.100},
            {"id":"buendia",       "nombre":"Buendía",        "buscar":["buendia","buendía"],  "rio":"Guadiela",   "municipio":"Buendía",            "cap":1651.0,"lat":40.391,"lon":-2.718},
        ]
    },
    "guadalajara": {
        "nombre": "Guadalajara", "comunidad": "Castilla-La Mancha",
        "embalses": [
            {"id":"entrepeñas",    "nombre":"Entrepeñas",     "buscar":["entrepeñas","entrepenas"],"rio":"Tajo",   "municipio":"Sacedón",            "cap":802.6, "lat":40.545,"lon":-2.691},
            {"id":"alcorlo",       "nombre":"Alcorlo",        "buscar":["alcorlo"],           "rio":"Bornoba",     "municipio":"Cogolludo",          "cap":180.0, "lat":40.991,"lon":-3.060},
            {"id":"belena",        "nombre":"Beleña",         "buscar":["beleña","belena"],    "rio":"Sorbe",      "municipio":"Tamajón",            "cap":53.0,  "lat":40.844,"lon":-3.123},
            {"id":"el_vado",       "nombre":"El Vado",        "buscar":["el vado guad","vado guadal"],"rio":"Jarama","municipio":"Campillo de Ranas","cap":56.0,  "lat":40.918,"lon":-3.322},
            {"id":"la_tajera",     "nombre":"La Tajera",      "buscar":["tajera"],            "rio":"Tajuña",      "municipio":"Cifuentes",          "cap":69.0,  "lat":40.773,"lon":-2.621},
        ]
    },
    # ══ EXTREMADURA ══
    "badajoz": {
        "nombre": "Badajoz", "comunidad": "Extremadura",
        "embalses": [
            {"id":"la_serena",     "nombre":"La Serena",      "buscar":["la serena"],         "rio":"Zújar",       "municipio":"Zalamea de la Serena","cap":3219.0,"lat":38.857,"lon":-5.470},
            {"id":"cijara",        "nombre":"Cijara",         "buscar":["cijara"],            "rio":"Guadiana",    "municipio":"Herrera del Duque",  "cap":1505.0,"lat":39.303,"lon":-5.010},
            {"id":"orellana",      "nombre":"Orellana",       "buscar":["orellana"],          "rio":"Guadiana",    "municipio":"Orellana la Vieja",  "cap":806.0, "lat":39.025,"lon":-5.542},
            {"id":"garcia_sola",   "nombre":"García de Sola", "buscar":["garcia de sola","garcía de sola"],"rio":"Guadiana","municipio":"Herrera del Duque","cap":554.0,"lat":39.016,"lon":-5.325},
            {"id":"alange",        "nombre":"Alange",         "buscar":["alange"],            "rio":"Matachel",    "municipio":"Alange",             "cap":852.0, "lat":38.788,"lon":-6.255},
            {"id":"zujar",         "nombre":"Zújar",          "buscar":["zujar","zújar"],     "rio":"Zújar",       "municipio":"Capilla",            "cap":308.0, "lat":38.727,"lon":-5.205},
            {"id":"sierra_brava",  "nombre":"Sierra Brava",   "buscar":["sierra brava"],      "rio":"Pizarroso",   "municipio":"Zorita",             "cap":232.0, "lat":39.280,"lon":-5.731},
            {"id":"villalba",      "nombre":"Villalba",       "buscar":["villalba"],          "rio":"Guadajira",   "municipio":"Villalba de los Barros","cap":106.0,"lat":38.666,"lon":-6.444},
        ]
    },
    "caceres": {
        "nombre": "Cáceres", "comunidad": "Extremadura",
        "embalses": [
            {"id":"alcantara",     "nombre":"Alcántara",      "buscar":["alcantara","alcántara"],"rio":"Tajo",     "municipio":"Alcántara",          "cap":3162.0,"lat":39.728,"lon":-6.892},
            {"id":"gabriel_galan", "nombre":"Gabriel y Galán","buscar":["gabriel y galan","gabriel galan"],"rio":"Alagón","municipio":"Guijo de Granadilla","cap":925.0,"lat":40.218,"lon":-6.086},
            {"id":"valdecanas_cac","nombre":"Valdecañas",     "buscar":["valdecañas","valdecanas"],"rio":"Tajo",   "municipio":"Berrocalejo",        "cap":1446.0,"lat":39.817,"lon":-5.440},
            {"id":"cedillo",       "nombre":"Cedillo",        "buscar":["cedillo"],           "rio":"Tajo",        "municipio":"Cedillo",            "cap":260.0, "lat":39.617,"lon":-7.476},
            {"id":"torrejon_tajo", "nombre":"Torrejón (Tajo-Tiétar)","buscar":["torrejon tajo","torrejón tajo"],"rio":"Tajo","municipio":"Serrejón","cap":188.0,  "lat":39.821,"lon":-5.748},
            {"id":"borbollon",     "nombre":"Borbollón",      "buscar":["borbollon","borbollón"],"rio":"Árrago",   "municipio":"Moraleja",           "cap":109.0, "lat":40.141,"lon":-6.557},
        ]
    },
    # ══ MADRID ══
    "madrid": {
        "nombre": "Madrid", "comunidad": "Comunidad de Madrid",
        "embalses": [
            {"id":"el_atazar",     "nombre":"El Atazar",      "buscar":["atazar"],            "rio":"Lozoya",      "municipio":"Patones",            "cap":426.0, "lat":40.903,"lon":-3.578},
            {"id":"valmayor",      "nombre":"Valmayor",       "buscar":["valmayor"],          "rio":"Aulencia",    "municipio":"Valdemorillo",       "cap":124.0, "lat":40.535,"lon":-4.052},
            {"id":"san_juan",      "nombre":"San Juan",       "buscar":["san juan mad","san juan madrid"],"rio":"Alberche","municipio":"Pelayos de la Presa","cap":138.0,"lat":40.371,"lon":-4.355},
            {"id":"santillana",    "nombre":"Santillana",     "buscar":["santillana"],        "rio":"Manzanares",  "municipio":"Manzanares el Real", "cap":91.0,  "lat":40.727,"lon":-3.891},
            {"id":"el_vellon",     "nombre":"El Vellón",      "buscar":["vellon","vellón"],   "rio":"Guadalix",    "municipio":"El Vellón",          "cap":41.0,  "lat":40.724,"lon":-3.589},
            {"id":"puentes_viejas","nombre":"Puentes Viejas", "buscar":["puentes viejas"],    "rio":"Lozoya",      "municipio":"Mangirón",           "cap":53.0,  "lat":40.967,"lon":-3.663},
            {"id":"riosequillo",   "nombre":"Riosequillo",    "buscar":["riosequillo"],       "rio":"Lozoya",      "municipio":"Buitrago de Lozoya", "cap":50.0,  "lat":40.974,"lon":-3.519},
            {"id":"el_pardo",      "nombre":"El Pardo",       "buscar":["el pardo"],          "rio":"Manzanares",  "municipio":"El Pardo",           "cap":43.0,  "lat":40.541,"lon":-3.780},
            {"id":"pinilla",       "nombre":"Pinilla",        "buscar":["pinilla"],           "rio":"Lozoya",      "municipio":"Pinilla del Valle",  "cap":38.0,  "lat":40.990,"lon":-3.685},
            {"id":"el_villar",     "nombre":"El Villar",      "buscar":["el villar mad","el villar lo"],"rio":"Lozoya","municipio":"Buitrago de Lozoya","cap":23.0,"lat":40.913,"lon":-3.631},
        ]
    },
    # ══ CYL ══
    "leon": {
        "nombre": "León", "comunidad": "Castilla y León",
        "embalses": [
            {"id":"riano",         "nombre":"Riaño",          "buscar":["riaño","riano"],     "rio":"Esla",        "municipio":"Riaño",              "cap":641.0, "lat":42.979,"lon":-5.005},
            {"id":"barrios_luna",  "nombre":"Barrios de Luna","buscar":["barrios de luna"],   "rio":"Luna",        "municipio":"Los Barrios de Luna","cap":308.0, "lat":42.850,"lon":-5.880},
            {"id":"barcena",       "nombre":"Bárcena",        "buscar":["barcena","bárcena"], "rio":"Sil",         "municipio":"Cubillos del Sil",   "cap":340.9, "lat":42.587,"lon":-6.576},
            {"id":"porma",         "nombre":"Porma",          "buscar":["porma"],             "rio":"Porma",       "municipio":"Valdehuesa",         "cap":318.0, "lat":43.009,"lon":-5.287},
        ]
    },
    "burgos": {
        "nombre": "Burgos", "comunidad": "Castilla y León",
        "embalses": [
            {"id":"uzquiza",       "nombre":"Uzquiza",        "buscar":["uzquiza"],           "rio":"Arlanzón",    "municipio":"Uzquiza",            "cap":71.0,  "lat":42.261,"lon":-3.576},
            {"id":"ordunte",       "nombre":"Ordunte",        "buscar":["ordunte"],           "rio":"Ordunte",     "municipio":"Valle de Mena",      "cap":22.0,  "lat":43.097,"lon":-3.335},
        ]
    },
    "palencia": {
        "nombre": "Palencia", "comunidad": "Castilla y León",
        "embalses": [
            {"id":"aguilar",       "nombre":"Aguilar de Campoo","buscar":["aguilar de campoo"],"rio":"Pisuerga",   "municipio":"Aguilar de Campoo",  "cap":247.0, "lat":42.795,"lon":-4.266},
            {"id":"compuerto",     "nombre":"Compuerto",      "buscar":["compuerto"],         "rio":"Carrión",     "municipio":"Velilla del R.Carrión","cap":95.0, "lat":42.905,"lon":-4.753},
            {"id":"camporredondo", "nombre":"Camporredondo",  "buscar":["camporredondo"],     "rio":"Carrión",     "municipio":"Alba de los Cardaños","cap":70.0,  "lat":42.980,"lon":-4.730},
        ]
    },
    "salamanca": {
        "nombre": "Salamanca", "comunidad": "Castilla y León",
        "embalses": [
            {"id":"almendra",      "nombre":"Almendra",       "buscar":["almendra"],          "rio":"Tormes",      "municipio":"Almendra",           "cap":3585.0,"lat":41.268,"lon":-6.343},
            {"id":"aldeadavila",   "nombre":"Aldeadávila",    "buscar":["aldeadavila","aldeadávila"],"rio":"Duero","municipio":"Aldeadávila",        "cap":114.3, "lat":41.225,"lon":-6.773},
            {"id":"saucelle",      "nombre":"Saucelle",       "buscar":["saucelle"],          "rio":"Duero",       "municipio":"Saucelle",           "cap":181.0, "lat":41.026,"lon":-6.747},
        ]
    },
    "zamora": {
        "nombre": "Zamora", "comunidad": "Castilla y León",
        "embalses": [
            {"id":"ricobayo",      "nombre":"Ricobayo",       "buscar":["ricobayo"],          "rio":"Esla",        "municipio":"Ricobayo",           "cap":1145.0,"lat":41.725,"lon":-5.850},
            {"id":"valparaiso",    "nombre":"Valparaíso",     "buscar":["valparaiso","valparaíso"],"rio":"Tera",   "municipio":"Palacios de Sanabria","cap":162.0,"lat":42.034,"lon":-6.527},
            {"id":"cernadilla",    "nombre":"Cernadilla",     "buscar":["cernadilla"],        "rio":"Tera",        "municipio":"Cernadilla",         "cap":255.0, "lat":42.026,"lon":-6.612},
        ]
    },
    "avila": {
        "nombre": "Ávila", "comunidad": "Castilla y León",
        "embalses": [
            {"id":"burguillo",     "nombre":"El Burguillo",   "buscar":["burguillo"],         "rio":"Alberche",    "municipio":"El Tiemblo",         "cap":197.7, "lat":40.413,"lon":-4.650},
            {"id":"castro_cogotas","nombre":"Castro Cogotas", "buscar":["castro de las cogotas","cogotas"],"rio":"Adaja","municipio":"Las Berlanas", "cap":59.0,  "lat":40.728,"lon":-4.895},
        ]
    },
    "segovia": {
        "nombre": "Segovia", "comunidad": "Castilla y León",
        "embalses": [
            {"id":"linares_arroyo","nombre":"Linares del Arroyo","buscar":["linares del arroyo","linares arroyo"],"rio":"Riaza","municipio":"Cerezo de Arriba","cap":59.0,"lat":41.348,"lon":-3.467},
            {"id":"burgomillodo",  "nombre":"Burgomillodo",   "buscar":["burgomillodo"],      "rio":"Duratón",     "municipio":"Sepúlveda",          "cap":13.7,  "lat":41.302,"lon":-3.556},
        ]
    },
    "soria": {
        "nombre": "Soria", "comunidad": "Castilla y León",
        "embalses": [
            {"id":"cuerda_pozo",   "nombre":"La Cuerda del Pozo","buscar":["cuerda del pozo","cuerda pozo"],"rio":"Duero","municipio":"Vinuesa",     "cap":249.0, "lat":41.966,"lon":-2.766},
        ]
    },
    # ══ ARAGÓN ══
    "zaragoza": {
        "nombre": "Zaragoza", "comunidad": "Aragón",
        "embalses": [
            {"id":"mequinenza",    "nombre":"Mequinenza",     "buscar":["mequinenza"],        "rio":"Ebro",        "municipio":"Mequinenza",         "cap":1373.0,"lat":41.381,"lon":0.276},
            {"id":"tranquera",     "nombre":"Tranquera",      "buscar":["tranquera"],         "rio":"Piedra",      "municipio":"La Tranquera",       "cap":81.6,  "lat":41.246,"lon":-1.844},
            {"id":"caspe",         "nombre":"Caspe",          "buscar":["caspe"],             "rio":"Guadalope",   "municipio":"Caspe",              "cap":82.0,  "lat":41.130,"lon":-0.175},
            {"id":"la_loteta",     "nombre":"La Loteta",      "buscar":["loteta"],            "rio":"Canal Imperial","municipio":"Gallur",           "cap":104.0, "lat":41.882,"lon":-1.316},
            {"id":"mularroya",     "nombre":"Mularroya",      "buscar":["mularroya"],         "rio":"Grío",        "municipio":"Mularroya",          "cap":96.9,  "lat":41.340,"lon":-1.544},
        ]
    },
    "huesca": {
        "nombre": "Huesca", "comunidad": "Aragón",
        "embalses": [
            {"id":"canelles",      "nombre":"Canelles",       "buscar":["canelles"],          "rio":"N.Ribagorzana","municipio":"Arén",              "cap":679.0, "lat":42.052,"lon":0.627},
            {"id":"mediano",       "nombre":"Mediano",        "buscar":["mediano"],           "rio":"Cinca",       "municipio":"Mediano",            "cap":436.0, "lat":42.269,"lon":0.146},
            {"id":"el_grado",      "nombre":"El Grado",       "buscar":["el grado"],          "rio":"Cinca",       "municipio":"El Grado",           "cap":399.0, "lat":42.136,"lon":0.171},
            {"id":"sotonera",      "nombre":"Sotonera",       "buscar":["sotonera"],          "rio":"Sotón",       "municipio":"Gurrea de Gállego",  "cap":189.0, "lat":42.136,"lon":-0.678},
            {"id":"santa_ana",     "nombre":"Santa Ana",      "buscar":["santa ana"],         "rio":"N.Ribagorzana","municipio":"Arén",              "cap":236.0, "lat":42.083,"lon":0.642},
            {"id":"escales",       "nombre":"Escales",        "buscar":["escales"],           "rio":"N.Ribagorzana","municipio":"Sopeira",           "cap":152.0, "lat":42.186,"lon":0.669},
            {"id":"barasona",      "nombre":"Barasona",       "buscar":["barasona"],          "rio":"Ésera",       "municipio":"Graus",              "cap":92.0,  "lat":42.186,"lon":0.411},
        ]
    },
    "teruel": {
        "nombre": "Teruel", "comunidad": "Aragón",
        "embalses": [
            {"id":"canon_santolea","nombre":"Cañon-Santolea", "buscar":["canon santolea","cañon santolea","santolea"],"rio":"Guadalope","municipio":"Castellote","cap":82.0,"lat":40.750,"lon":-0.300},
            {"id":"cueva_foradada","nombre":"Cueva Foradada", "buscar":["cueva foradada"],    "rio":"Martín",      "municipio":"Oliete",             "cap":29.0,  "lat":40.996,"lon":-0.628},
            {"id":"arquillo",      "nombre":"Arquillo de San Blas","buscar":["arquillo"],     "rio":"Turia",       "municipio":"Gea de Albarracín",  "cap":21.0,  "lat":40.406,"lon":-1.382},
        ]
    },
    # ══ C.VALENCIANA ══
    "castellon": {
        "nombre": "Castellón", "comunidad": "Comunitat Valenciana",
        "embalses": [
            {"id":"arenos",        "nombre":"Arenós",         "buscar":["arenos","arenós"],   "rio":"Mijares",     "municipio":"Puebla de Arenoso",  "cap":137.0, "lat":40.109,"lon":-0.590},
            {"id":"sichar",        "nombre":"Sichar",         "buscar":["sichar"],            "rio":"Mijares",     "municipio":"Espadilla",          "cap":49.0,  "lat":39.971,"lon":-0.446},
        ]
    },
    "valencia": {
        "nombre": "Valencia", "comunidad": "Comunitat Valenciana",
        "embalses": [
            {"id":"tous",          "nombre":"Tous",           "buscar":["tous"],              "rio":"Júcar",       "municipio":"Tous",               "cap":379.0, "lat":39.194,"lon":-0.832},
            {"id":"benageber",     "nombre":"Benagéber",      "buscar":["benageber","benagéber"],"rio":"Turia",    "municipio":"Benagéber",          "cap":228.0, "lat":39.681,"lon":-1.086},
            {"id":"cortes_ii",     "nombre":"Cortes II",      "buscar":["cortes ii"],         "rio":"Júcar",       "municipio":"Cortes de Pallás",   "cap":118.0, "lat":39.194,"lon":-1.195},
            {"id":"escalona",      "nombre":"Escalona",       "buscar":["escalona"],          "rio":"Cabriel",     "municipio":"Requena",            "cap":99.0,  "lat":39.440,"lon":-1.370},
        ]
    },
    "alicante": {
        "nombre": "Alicante", "comunidad": "Comunitat Valenciana",
        "embalses": [
            {"id":"la_pedrera",    "nombre":"La Pedrera",     "buscar":["la pedrera","pedrera"],"rio":"Alcorisa",  "municipio":"Orihuela",           "cap":246.0, "lat":37.990,"lon":-0.900},
            {"id":"beniarres",     "nombre":"Beniarrés",      "buscar":["beniarres","beniarrés"],"rio":"Serpis",   "municipio":"Beniarrés",          "cap":27.0,  "lat":38.781,"lon":-0.330},
        ]
    },
    # ══ CATALUÑA ══
    "lleida": {
        "nombre": "Lleida", "comunidad": "Cataluña",
        "embalses": [
            {"id":"rialb",         "nombre":"Rialb",          "buscar":["rialb"],             "rio":"Segre",       "municipio":"Peramola",           "cap":403.0, "lat":42.013,"lon":1.281},
            {"id":"talarn",        "nombre":"Talarn",         "buscar":["talarn"],            "rio":"N.Pallaresa", "municipio":"Tremp",              "cap":227.0, "lat":42.048,"lon":0.907},
            {"id":"camarasa",      "nombre":"Camarasa",       "buscar":["camarasa"],          "rio":"N.Pallaresa", "municipio":"Camarasa",           "cap":163.0, "lat":41.891,"lon":0.936},
            {"id":"oliana",        "nombre":"Oliana",         "buscar":["oliana"],            "rio":"Segre",       "municipio":"Oliana",             "cap":84.0,  "lat":42.073,"lon":1.371},
            {"id":"llosa_cavall",  "nombre":"La Llosa del Cavall","buscar":["llosa del cavall","llosa cavall"],"rio":"Cardener","municipio":"Navès", "cap":80.0,  "lat":42.052,"lon":1.689},
        ]
    },
    "girona": {
        "nombre": "Girona", "comunidad": "Cataluña",
        "embalses": [
            {"id":"susqueda",      "nombre":"Susqueda",       "buscar":["susqueda"],          "rio":"Ter",         "municipio":"Susqueda",           "cap":233.0, "lat":41.968,"lon":2.538},
            {"id":"boadella",      "nombre":"Boadella",       "buscar":["boadella"],          "rio":"La Muga",     "municipio":"Darnius",            "cap":61.0,  "lat":42.321,"lon":2.817},
        ]
    },
    "barcelona": {
        "nombre": "Barcelona", "comunidad": "Cataluña",
        "embalses": [
            {"id":"sau",           "nombre":"Sau",            "buscar":["sau"],               "rio":"Ter",         "municipio":"Tavèrnoles",         "cap":165.0, "lat":41.984,"lon":2.427},
            {"id":"la_baells",     "nombre":"La Baells",      "buscar":["baells"],            "rio":"Llobregat",   "municipio":"Cercs",              "cap":115.0, "lat":41.955,"lon":1.921},
        ]
    },
    "tarragona": {
        "nombre": "Tarragona", "comunidad": "Cataluña",
        "embalses": [
            {"id":"ribarroja_tar", "nombre":"Ribarroja",      "buscar":["ribarroja"],         "rio":"Ebro",        "municipio":"Riba-roja d'Ebre",   "cap":210.0, "lat":41.276,"lon":0.487},
        ]
    },
    # ══ GALICIA ══
    "a_coruna": {
        "nombre": "A Coruña", "comunidad": "Galicia",
        "embalses": [
            {"id":"eume",          "nombre":"Eume",           "buscar":["eume"],              "rio":"Eume",        "municipio":"As Pontes",          "cap":123.0, "lat":43.485,"lon":-7.953},
            {"id":"fervenza",      "nombre":"Fervenza",       "buscar":["fervenza"],          "rio":"Xallas",      "municipio":"Santa Comba",        "cap":103.0, "lat":42.984,"lon":-8.788},
            {"id":"cecebre",       "nombre":"Cecebre",        "buscar":["cecebre"],           "rio":"Mero",        "municipio":"Cambre",             "cap":22.0,  "lat":43.272,"lon":-8.243},
        ]
    },
    "lugo": {
        "nombre": "Lugo", "comunidad": "Galicia",
        "embalses": [
            {"id":"belesar",       "nombre":"Belesar",        "buscar":["belesar"],           "rio":"Miño",        "municipio":"Chantada",           "cap":655.0, "lat":42.600,"lon":-7.717},
            {"id":"los_peares",    "nombre":"Los Peares",     "buscar":["peares"],            "rio":"Miño",        "municipio":"O Carballiño",       "cap":182.0, "lat":42.431,"lon":-7.895},
        ]
    },
    "ourense": {
        "nombre": "Ourense", "comunidad": "Galicia",
        "embalses": [
            {"id":"as_portas",     "nombre":"As Portas",      "buscar":["as portas","portas"], "rio":"Camba",      "municipio":"A Pobra de Trives",  "cap":535.8, "lat":42.337,"lon":-7.161},
            {"id":"bao",           "nombre":"Bao",            "buscar":["bao"],               "rio":"Bibei",       "municipio":"A Pobra de Trives",  "cap":238.1, "lat":42.260,"lon":-7.018},
            {"id":"santo_estevo",  "nombre":"Santo Estevo",   "buscar":["santo estevo"],      "rio":"Sil",         "municipio":"Nogueira de Ramuín", "cap":213.0, "lat":42.374,"lon":-7.617},
            {"id":"prada",         "nombre":"Prada",          "buscar":["prada"],             "rio":"Xares",       "municipio":"A Veiga",            "cap":122.0, "lat":42.232,"lon":-6.832},
        ]
    },
    "pontevedra": {
        "nombre": "Pontevedra", "comunidad": "Galicia",
        "embalses": [
            {"id":"portodemouros", "nombre":"Portodemouros",  "buscar":["portodemouros"],     "rio":"Ulla",        "municipio":"Vila de Cruces",     "cap":297.0, "lat":42.803,"lon":-8.261},
            {"id":"eiras",         "nombre":"Eiras",          "buscar":["eiras"],             "rio":"Oitavén",     "municipio":"Fornelos de Montes", "cap":22.0,  "lat":42.380,"lon":-8.433},
        ]
    },
    # ══ PAÍS VASCO ══
    "alava": {
        "nombre": "Álava", "comunidad": "País Vasco",
        "embalses": [
            {"id":"ullivarri",     "nombre":"Ullíbarri-Gamboa","buscar":["ullibarri","ullíbarri"],"rio":"Zadorra",  "municipio":"Ullíbarri-Gamboa",  "cap":146.0, "lat":42.840,"lon":-2.638},
            {"id":"urrunaga",      "nombre":"Urrunaga",       "buscar":["urrunaga"],          "rio":"Santa Engracia","municipio":"Legutiano",        "cap":72.0,  "lat":42.977,"lon":-2.671},
        ]
    },
    "gipuzkoa": {
        "nombre": "Guipúzcoa", "comunidad": "País Vasco",
        "embalses": [
            {"id":"urkulu",        "nombre":"Urkulu",         "buscar":["urkulu"],            "rio":"Urkulu",      "municipio":"Oñati",              "cap":10.0,  "lat":43.044,"lon":-2.422},
            {"id":"ibai_eder",     "nombre":"Ibai-Eder",      "buscar":["ibai eder","ibai-eder"],"rio":"Urrestilla","municipio":"Azpeitia",          "cap":11.0,  "lat":43.180,"lon":-2.285},
        ]
    },
    # ══ NAVARRA ══
    "navarra": {
        "nombre": "Navarra", "comunidad": "Navarra",
        "embalses": [
            {"id":"yesa",          "nombre":"Yesa",           "buscar":["yesa"],              "rio":"Aragón",      "municipio":"Yesa",               "cap":447.0, "lat":42.618,"lon":-1.180},
            {"id":"itoiz",         "nombre":"Itoiz",          "buscar":["itoiz"],             "rio":"Irati",       "municipio":"Itoiz",              "cap":418.0, "lat":42.750,"lon":-1.435},
            {"id":"alloz",         "nombre":"Alloz",          "buscar":["alloz"],             "rio":"Salado",      "municipio":"Guesálaz",           "cap":65.0,  "lat":42.693,"lon":-1.950},
            {"id":"eugui",         "nombre":"Eugi",           "buscar":["eugui","eugi"],      "rio":"Arga",        "municipio":"Eugi",               "cap":22.0,  "lat":42.946,"lon":-1.581},
        ]
    },
    # ══ LA RIOJA ══
    "la_rioja": {
        "nombre": "La Rioja", "comunidad": "La Rioja",
        "embalses": [
            {"id":"mansilla",      "nombre":"Mansilla",       "buscar":["mansilla"],          "rio":"Najerilla",   "municipio":"Mansilla de la Sierra","cap":68.0, "lat":42.175,"lon":-2.905},
            {"id":"pajares",       "nombre":"Pajares",        "buscar":["pajares"],           "rio":"Piqueras",    "municipio":"San Pedro Manrique", "cap":35.0,  "lat":42.073,"lon":-2.424},
            {"id":"enciso",        "nombre":"Enciso",         "buscar":["enciso"],            "rio":"Cidacos",     "municipio":"Enciso",             "cap":46.0,  "lat":42.189,"lon":-2.355},
        ]
    },
    # ══ CANTABRIA ══
    "cantabria": {
        "nombre": "Cantabria", "comunidad": "Cantabria",
        "embalses": [
            {"id":"ebro_cant",     "nombre":"Del Ebro",       "buscar":["del ebro","embalse del ebro"],"rio":"Ebro","municipio":"Arija",             "cap":540.0, "lat":42.973,"lon":-4.007},
            {"id":"alsa_mediajo",  "nombre":"Alsa-Mediajo",   "buscar":["alsa","mediajo"],    "rio":"Torina",      "municipio":"Los Corrales de Buelna","cap":22.0,"lat":43.227,"lon":-4.028},
        ]
    },
    # ══ ASTURIAS ══
    "asturias": {
        "nombre": "Asturias", "comunidad": "Asturias",
        "embalses": [
            {"id":"salime",        "nombre":"Salime",         "buscar":["salime"],            "rio":"Navia",       "municipio":"Grandas de Salime",  "cap":237.8, "lat":43.261,"lon":-6.797},
            {"id":"doiras",        "nombre":"Doiras",         "rio":"Navia",                  "buscar":["doiras"], "municipio":"Boal",               "cap":100.0, "lat":43.342,"lon":-6.791},
            {"id":"tanes",         "nombre":"Tanes",          "buscar":["tanes"],             "rio":"Nalón",       "municipio":"Caso",               "cap":33.0,  "lat":43.204,"lon":-5.489},
        ]
    },
}

# ── FUNCIÓN DE CRUCE ───────────────────────────────────────────────────────────

def buscar_en_excel(embalse: dict, datos_excel: dict) -> tuple:
    """
    Busca los datos de un embalse en el diccionario del Excel.
    Devuelve (volumen_hm3, pct) o (None, None) si no encuentra.
    """
    for termino in embalse["buscar"]:
        termino_n = norm(termino)
        for clave_excel, vals in datos_excel.items():
            clave_n = norm(clave_excel)
            if termino_n in clave_n or clave_n in termino_n:
                return vals.get("vol"), vals.get("pct")
    return None, None

def calcular_estado(pct):
    if pct is None:  return "#888888", "Sin datos"
    p = min(pct, 100)
    if p < 20:       return "#CC2200", "Crítico"
    if p < 40:       return "#FF8822", "Bajo"
    if p < 60:       return "#FFCC44", "Moderado"
    if p < 80:       return "#44AA66", "Bueno"
    return "#0066CC", "Muy bueno"

# ── GENERACIÓN DE JSONS ────────────────────────────────────────────────────────

def procesar_todo(datos_excel: dict, fecha_boletin: date):
    os.makedirs("docs/embalses", exist_ok=True)
    fecha_str = fecha_boletin.strftime("%d/%m/%Y")
    fecha_iso = datetime.now().isoformat()
    resumen_nacional = []

    for id_prov, prov in PROVINCIAS.items():
        lista_embalses = []
        total_vol = total_cap = 0.0
        tiene_datos = False

        for emb in prov["embalses"]:
            vol, pct = buscar_en_excel(emb, datos_excel)

            # Si el Excel tiene volumen pero no porcentaje, calcularlo
            if vol is not None and pct is None and emb["cap"] > 0:
                pct = round(vol / emb["cap"] * 100, 1)
            # Si tiene porcentaje pero no volumen, calcularlo
            if pct is not None and vol is None:
                vol = round(emb["cap"] * pct / 100, 1)

            if pct is not None:
                tiene_datos = True
                total_vol += vol or 0.0
                total_cap += emb["cap"]
                pct = round(min(pct, 100), 1)
            else:
                total_cap += emb["cap"]

            color, etiqueta = calcular_estado(pct)
            lista_embalses.append({
                "id":           emb["id"],
                "nombre":       emb["nombre"],
                "rio":          emb.get("rio", "—"),
                "municipio":    emb.get("municipio", "—"),
                "provincia":    prov["nombre"],
                "lat":          emb["lat"],
                "lon":          emb["lon"],
                "capacidad_hm3":emb["cap"],
                "volumen_hm3":  round(vol, 1) if vol is not None else None,
                "pct":          pct,
                "color":        color,
                "etiqueta":     etiqueta,
            })

        pct_media = round(total_vol / total_cap * 100, 1) if total_cap > 0 and tiene_datos else None
        color_m, etq_m = calcular_estado(pct_media)

        json_prov = {
            "ultima_actualizacion": fecha_iso,
            "fecha_legible":        fecha_str,
            "provincia":            prov["nombre"],
            "comunidad":            prov["comunidad"],
            "total_embalses":       len(lista_embalses),
            "capacidad_total_hm3":  round(total_cap, 1),
            "volumen_total_hm3":    round(total_vol, 1),
            "pct_media":            pct_media,
            "color":                color_m,
            "etiqueta":             etq_m,
            "fuente":               f"Boletín Hidrológico Semanal — MITECO ({fecha_str})",
            "embalses":             lista_embalses,
        }

        ruta = f"docs/embalses/{id_prov}.json"
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(json_prov, f, ensure_ascii=False, indent=2)

        resumen_nacional.append({
            "id":               id_prov,
            "nombre":           prov["nombre"],
            "comunidad":        prov["comunidad"],
            "pct":              pct_media,
            "color":            color_m,
            "etiqueta":         etq_m,
            "url_detalle":      f"embalses/{id_prov}.html",
            "datos_disponibles":tiene_datos,
        })

        estado_str = f"{pct_media}%" if pct_media else "sin datos"
        print(f"  ✓ {prov['nombre']:16s} → {estado_str:8s}  "
              f"({round(total_vol,0):.0f}/{round(total_cap,0):.0f} hm³)  "
              f"[{len(lista_embalses)} embalses]")

    # Nacional
    nacional = {
        "ultima_actualizacion": fecha_iso,
        "fecha_legible":        fecha_str,
        "fuente":               f"Boletín Hidrológico Semanal — MITECO ({fecha_str})",
        "provincias":           resumen_nacional,
        "comunidades":          resumen_nacional,
    }
    with open("docs/embalses_nacional.json", "w", encoding="utf-8") as f:
        json.dump(nacional, f, ensure_ascii=False, indent=2)
    print(f"\n✓ docs/embalses_nacional.json actualizado")
    print(f"✓ {len(PROVINCIAS)} provincias procesadas")

# ── MAIN ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("Boletín Hidrológico Semanal — MITECO")
    print("=" * 60)

    print("\n1. Descargando Excel del boletín...")
    contenido, fecha_boletin = descargar_excel(max_semanas_atras=5)

    if contenido is None:
        print("\n⚠️  No se pudo descargar el Excel.")
        print("   Posibles causas:")
        print("   - El MITECO no ha publicado aún el boletín esta semana")
        print("   - Cambio en la URL del servidor")
        print("   Revisando la última semana disponible manualmente en:")
        print("   https://www.miteco.gob.es/es/agua/temas/evaluacion-de-los-recursos-hidricos/boletin-hidrologico.html")
        exit(1)

    print("\n2. Leyendo datos del Excel...")
    datos_excel = leer_excel(contenido)

    if not datos_excel:
        print("⚠️  El Excel no contiene datos procesables.")
        exit(1)

    print(f"\n3. Generando JSONs para {len(PROVINCIAS)} provincias...")
    procesar_todo(datos_excel, fecha_boletin)

    print("\n✓ Proceso completado correctamente.")
